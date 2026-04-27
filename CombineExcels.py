"""
Combine per-page Excel files into a single Output.xlsx.

Pages/1/1.xlsx  — copied as-is (header + data)
Pages/N/N.xlsx  — data rows only (header skipped)

Usage:
    python CombineExcels.py
    python CombineExcels.py --input Pages --output Output.xlsx
"""

import os
import argparse
import shutil
from openpyxl import load_workbook


def main():
    parser = argparse.ArgumentParser(description="Combine per-page Excel files into one.")
    parser.add_argument("--input", default="Pages", help="Folder containing page sub-folders (default: Pages)")
    parser.add_argument("--output", default="Output.xlsx", help="Output file path (default: Output.xlsx)")
    args = parser.parse_args()

    pages = sorted(
        int(p) for p in os.listdir(args.input)
        if os.path.isdir(os.path.join(args.input, p)) and p.isdigit()
    )

    if not pages:
        print("Không tìm thấy trang nào.")
        return

    # Verify all files exist before starting
    missing = [p for p in pages if not os.path.exists(os.path.join(args.input, str(p), f"{p}.xlsx"))]
    if missing:
        print(f"Thiếu file cho các trang: {missing}. Hãy chạy pipeline trước.")
        return

    first_path = os.path.join(args.input, str(pages[0]), f"{pages[0]}.xlsx")
    shutil.copy2(first_path, args.output)
    print(f"Trang {pages[0]} — sao chép làm nền -> {args.output}")

    if len(pages) == 1:
        print("Chỉ có một trang, hoàn thành.")
        return

    wb_out = load_workbook(args.output)
    ws_out = wb_out.active

    for page in pages[1:]:
        path = os.path.join(args.input, str(page), f"{page}.xlsx")
        wb_in = load_workbook(path, data_only=True)
        ws_in = wb_in.active

        row_count = 0
        for row in ws_in.iter_rows(min_row=2, values_only=True):  # skip header
            ws_out.append(list(row))
            row_count += 1

        print(f"Trang {page} — thêm {row_count} dòng")

    wb_out.save(args.output)
    total = ws_out.max_row - 1  # exclude header
    print(f"\nHoàn thành. {total} dòng dữ liệu -> {args.output}")


if __name__ == "__main__":
    main()
