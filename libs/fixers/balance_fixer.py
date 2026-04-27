"""
Fixer that validates each row's balance against the accounting equation:

    Balance = Previous Balance - Debit + Credit

When the equation does not hold, the fixer attempts to recover by removing
one digit at a time from the OCR'd balance and checking whether the result
matches the expected value. If a single-digit removal produces a match, the
cell is corrected in place and the row is marked fixed. If no removal
produces a match, the row is marked as an error only.

Validation is skipped for a row if any of the three key values (previous
balance, current debit, current credit) are not integers, which means the
formatter could not parse them.

For the first data row the previous balance is taken from the last row of
the preceding page's refined Excel file (e.g. Pages/399/399.xlsx when
processing page 400).  Pass raw_path so the fixer can locate that file.
"""
import os

from openpyxl import load_workbook


class NegativeBalanceError(Exception):
    """Raised after a page is processed when one or more expected balances are negative."""


def _try_remove_one_digit(balance: int, expected: int):
    """Return (corrected_int, position, removed_digit) if removing exactly one
    digit from balance yields expected, otherwise return None."""
    s = str(balance)
    for i, digit in enumerate(s):
        candidate_str = s[:i] + s[i + 1:]
        if not candidate_str:
            continue
        # Avoid leading zeros creating a shorter number unintentionally
        candidate = int(candidate_str)
        if candidate == expected:
            return candidate, i, digit
    return None


def _apply_balance_fix(ws, row, col, balance, expected, row_errors, row_fixes):
    """Apply the standard mismatch fix (remove-digit or overwrite) for one row."""
    fix = _try_remove_one_digit(balance, expected)
    if fix:
        corrected, pos, digit = fix
        ws.cell(row=row, column=col).value = corrected
        row_errors.setdefault(row, {})[col] = (
            f"Số dư không khớp: kỳ vọng {expected:,}, OCR đọc được {balance:,}"
        )
        row_fixes.setdefault(row, []).append(
            f"Số dư: đã xóa chữ số thừa '{digit}' tại vị trí {pos} "
            f"(OCR đọc được {balance:,}, đã sửa thành {corrected:,})"
        )
    else:
        ws.cell(row=row, column=col).value = expected
        row_errors.setdefault(row, {})[col] = (
            f"Số dư không khớp: kỳ vọng {expected:,}, OCR đọc được {balance:,}"
        )
        row_fixes.setdefault(row, []).append(
            f"Khôi phục số dư bằng cách ghi đè giá trị kỳ vọng {expected:,} "
            f"(OCR đọc được {balance:,}, đã sửa thành {expected:,})"
        )


def _prev_page_last_balance(raw_path, col):
    """Return the last integer balance from the previous page's refined Excel, or None."""
    page_dir = os.path.dirname(os.path.abspath(raw_path))
    pages_dir = os.path.dirname(page_dir)
    try:
        page_num = int(os.path.basename(page_dir))
    except ValueError:
        return None

    prev_excel = os.path.join(pages_dir, str(page_num - 1), f"{page_num - 1}.xlsx")
    if not os.path.exists(prev_excel):
        return None

    try:
        prev_wb = load_workbook(prev_excel, read_only=True, data_only=True)
        prev_ws = prev_wb.active
        prev_balance = None
        for r in range(prev_ws.max_row, 1, -1):  # skip header row 1
            v = prev_ws.cell(row=r, column=col).value
            if isinstance(v, int):
                prev_balance = v
                break
        prev_wb.close()
        return prev_balance
    except Exception:
        return None


def _page_number(raw_path):
    """Return the page number derived from raw_path's parent directory, or None."""
    if raw_path is None:
        return None
    try:
        return int(os.path.basename(os.path.dirname(os.path.abspath(raw_path))))
    except ValueError:
        return None


def fix_balance(ws, row_fixes, row_errors, debit_col, credit_col, col, raw_path=None):
    """
    Args:
        debit_col:  1-based column index of the Debit column.
        credit_col: 1-based column index of the Credit column.
        col:        1-based column index of the Balance column.
        raw_path:   Path to the current page's raw Excel file.  When supplied,
                    the first data row is validated against the previous page.
    """
    DATA_START = 2
    page = _page_number(raw_path)
    negative_rows = []  # collect all negative-expected rows; raise after processing

    # --- First data row: validate against previous page's last balance ---
    if raw_path is not None:
        prev_balance = _prev_page_last_balance(raw_path, col)
        if prev_balance is not None:
            row = DATA_START
            debit  = ws.cell(row=row, column=debit_col).value
            credit = ws.cell(row=row, column=credit_col).value
            balance = ws.cell(row=row, column=col).value

            if isinstance(balance, int):
                if debit == "":
                    debit = None
                if credit == "":
                    credit = None
                if (debit is None or isinstance(debit, int)) and \
                   (credit is None or isinstance(credit, int)):
                    expected = prev_balance - (debit or 0) + (credit or 0)
                    if expected < 0:
                        row_errors.setdefault(row, {})[col] = (
                            f"Số dư kỳ vọng âm ({expected:,}) — cần sửa thủ công"
                        )
                        negative_rows.append((row, expected))
                    elif expected != balance:
                        _apply_balance_fix(ws, row, col, balance, expected, row_errors, row_fixes)

    # --- Remaining rows: validate each row against the one above ---
    for row in range(DATA_START + 1, ws.max_row + 1):
        prev_balance = ws.cell(row=row - 1, column=col).value
        debit        = ws.cell(row=row, column=debit_col).value
        credit       = ws.cell(row=row, column=credit_col).value
        balance      = ws.cell(row=row, column=col).value

        # Require prev_balance and balance to be formatted integers
        if not isinstance(prev_balance, int) or not isinstance(balance, int):
            continue
        # Blank OCR cells arrive as "" — treat the same as None (i.e. zero)
        if debit == "":
            debit = None
        if credit == "":
            credit = None
        # If debit or credit failed formatting (non-empty string), skip
        if debit is not None and not isinstance(debit, int):
            continue
        if credit is not None and not isinstance(credit, int):
            continue

        expected = prev_balance - (debit or 0) + (credit or 0)
        if expected < 0:
            row_errors.setdefault(row, {})[col] = (
                f"Số dư kỳ vọng âm ({expected:,}) — cần sửa thủ công"
            )
            negative_rows.append((row, expected))
            continue
        if expected == balance:
            continue

        _apply_balance_fix(ws, row, col, balance, expected, row_errors, row_fixes)

    if negative_rows:
        detail = ", ".join(f"dòng {r} ({v:,})" for r, v in negative_rows)
        raise NegativeBalanceError(
            f"Trang {page}: số dư kỳ vọng âm tại {detail}. "
            "File đã được lưu — vui lòng kiểm tra và sửa thủ công."
        )
