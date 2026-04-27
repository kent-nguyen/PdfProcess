"""
Fixer for sequential serial-number (STT) columns.
"""
import os
from openpyxl import load_workbook


def _int_or_none(value):
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def _prev_page_last_stt(raw_path, col):
    """Return the last integer STT from the previous page's refined Excel, or None."""
    page_dir = os.path.dirname(os.path.abspath(raw_path))
    pages_dir = os.path.dirname(page_dir)
    try:
        page_num = int(os.path.basename(page_dir))
    except ValueError:
        return None

    prev_excel = os.path.join(pages_dir, str(page_num - 1), f"{page_num - 1}.xlsx")
    if not os.path.exists(prev_excel):
        return None

    try:
        prev_wb = load_workbook(prev_excel, read_only=True, data_only=True)
        prev_ws = prev_wb.active
        last_stt = None
        for r in range(prev_ws.max_row, 1, -1):  # skip header row 1
            v = prev_ws.cell(row=r, column=col).value
            if isinstance(v, int):
                last_stt = v
                break
        prev_wb.close()
        return last_stt
    except Exception:
        return None


def fix_stt(ws, row_fixes, row_errors, col, raw_path=None, **_):
    """
    Fix empty or malformed cells in a sequential serial-number column.

    Reads row_errors (populated by the formatter) to detect cells that had
    bad values (e.g. 'S2') and includes the original value in fix notes.

    Strategy:
      - Fill by incrementing from the previous known value.
      - If the result disagrees with the next known value, append a warning.
      - If no previous value exists, decrement back from the next known value.
      - If neither neighbour is known, flag for manual review.

    Args:
        col: 1-based column index of the STT column.
    """
    DATA_START = 2  # row 1 is the header

    # --- First row: fix against previous page's last STT ---
    if raw_path is not None:
        prev_last = _prev_page_last_stt(raw_path, col)
        if prev_last is not None:
            expected = prev_last + 1
            cell = ws.cell(row=DATA_START, column=col)
            v = _int_or_none(cell.value)
            if v != expected:
                cell.value = expected
                row_fixes.setdefault(DATA_START, []).append(
                    f"STT: giá trị là {v!r}, đã sửa thành {expected} (tiếp theo từ trang trước {prev_last})"
                )

    known = {}
    for row in range(DATA_START, ws.max_row + 1):
        v = _int_or_none(ws.cell(row=row, column=col).value)
        if v is not None:
            known[row] = v

    empty = [r for r in range(DATA_START, ws.max_row + 1) if r not in known]

    if empty:
        groups = []
        group = [empty[0]]
        for r in empty[1:]:
            if r == group[-1] + 1:
                group.append(r)
            else:
                groups.append(group)
                group = [r]
        groups.append(group)

    for group in (groups if empty else []):
        first, last = group[0], group[-1]
        prev_val = known.get(first - 1)
        next_val = known.get(last + 1)

        if prev_val is not None:
            filled = [prev_val + i + 1 for i in range(len(group))]
            for row, val in zip(group, filled):
                ws.cell(row=row, column=col).value = val
                bad_val = row_errors.get(row, {}).get(col)
                origin = f"bad value '{bad_val}'" if bad_val else "empty"
                note = f"STT: giá trị là {origin}, đã điền là {val} (tăng dần từ {prev_val})"
                if next_val is not None and filled[-1] + 1 != next_val:
                    note += f" [CẢNH BÁO: chuỗi bị đứt — giá trị tiếp theo đã biết là {next_val}]"
                row_fixes.setdefault(row, []).append(note)

        elif next_val is not None:
            filled = [next_val - (len(group) - i) for i in range(len(group))]
            for row, val in zip(group, filled):
                ws.cell(row=row, column=col).value = val
                bad_val = row_errors.get(row, {}).get(col)
                origin = f"bad value '{bad_val}'" if bad_val else "empty"
                row_fixes.setdefault(row, []).append(
                    f"STT: giá trị là {origin}, đã điền là {val} (giảm dần từ {next_val})"
                )

        else:
            for row in group:
                bad_val = row_errors.get(row, {}).get(col)
                origin = f"bad value '{bad_val}'" if bad_val else "empty"
                row_fixes.setdefault(row, []).append(
                    f"STT: giá trị là {origin}, không có giá trị liền kề để nội suy — cần kiểm tra thủ công"
                )

    # Second pass: verify sequence continuity on all rows (catches valid-integer
    # misreads like 274 → 214 that the formatter cannot detect).
    # Corrects purely based on the previous row — cascading is intentional so a
    # run of consistently misread values (e.g. 214,215,216… instead of 274,275,276…)
    # is fully repaired in one pass.
    for row in range(DATA_START, ws.max_row + 1):
        v = _int_or_none(ws.cell(row=row, column=col).value)
        if v is None:
            continue
        prev_v = _int_or_none(ws.cell(row=row - 1, column=col).value) if row > DATA_START else None
        if prev_v is None or v == prev_v + 1:
            continue
        expected = prev_v + 1
        ws.cell(row=row, column=col).value = expected
        row_fixes.setdefault(row, []).append(
            f"STT: giá trị là {v}, đã sửa thành {expected} (sai thứ tự sau {prev_v})"
        )
