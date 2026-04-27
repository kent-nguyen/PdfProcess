import os
import re
import argparse
import importlib
import warnings
warnings.filterwarnings("ignore", message=".*pin_memory.*")
import easyocr
from openpyxl import Workbook


def ocr_cell(reader, image_path, allowlist=None):
    kwargs = {"allowlist": allowlist} if allowlist else {}
    result = reader.readtext(image_path, **kwargs)
    if not result:
        return ""
    # Sort top-to-bottom then left-to-right. Cluster detections that share roughly the
    # same vertical band (within half the average box height) into the same row so that
    # numbers like "35,700,000.00" — where EasyOCR splits "35" and "700,000.00" into
    # two boxes whose Y-tops differ by a few pixels — still sort left-to-right.
    heights = [abs(r[0][2][1] - r[0][0][1]) for r in result]
    avg_h = sum(heights) / len(heights) if heights else 15
    row_band = max(1, int(avg_h * 0.5))
    result = sorted(result, key=lambda r: (r[0][0][1] // row_band, r[0][0][0]))
    return " ".join(item[1] for item in result).strip()


def cells_to_excel(page_dir, reader, column_allowlists):
    cells_dir = os.path.join(page_dir, "Cells")
    if not os.path.isdir(cells_dir):
        print(f"  Không tìm thấy thư mục Cells trong {page_dir}, bỏ qua.")
        return

    pattern = re.compile(r"row(\d+)_col(\d+)\.png$")
    cell_files = {}
    for fname in os.listdir(cells_dir):
        m = pattern.match(fname)
        if m:
            cell_files[(int(m.group(1)), int(m.group(2)))] = os.path.join(cells_dir, fname)

    if not cell_files:
        print("  Không tìm thấy ảnh ô nào.")
        return

    # Detect a spurious empty first column: OCR the header cell (row 0, col 0)
    # without an allowlist. If it returns blank, all real data columns are shifted
    # right by 1 in the image, so we shift left by 1 when writing to Excel.
    col_offset = 0
    first_cell_path = cell_files.get((0, 0))
    if first_cell_path and ocr_cell(reader, first_cell_path) == "":
        col_offset = 1
        print("  Phát hiện cột đầu tiên trống — bỏ qua cột 0, dịch trái 1 cột")

    wb = Workbook()
    ws = wb.active

    max_row = max(r for r, _ in cell_files)
    current_row = None
    for (row, col) in sorted(cell_files):
        if col < col_offset:
            continue  # skip the spurious empty column
        if row != current_row:
            current_row = row
            print(f"  Dòng {row}/{max_row}", end="\r", flush=True)
        logical_col = col - col_offset
        allowlist = column_allowlists.get(logical_col) if row > 0 else None
        text = ocr_cell(reader, cell_files[(row, col)], allowlist=allowlist)
        ws.cell(row=row + 1, column=logical_col + 1, value=text)
    print(f"  Dòng {max_row}/{max_row} — hoàn thành")

    # Remove rows where the first 4 data columns are all empty (skip header row 1).
    rows_to_delete = []
    for ws_row in ws.iter_rows(min_row=2):
        if all((ws_row[c].value or "").strip() == "" for c in range(min(4, len(ws_row)))):
            rows_to_delete.append(ws_row[0].row)
    for r in reversed(rows_to_delete):
        ws.delete_rows(r)
    if rows_to_delete:
        print(f"  Đã xoá {len(rows_to_delete)} dòng trống: {rows_to_delete}")

    page_name = os.path.basename(page_dir)
    out_path = os.path.join(page_dir, f"raw_{page_name}.xlsx")
    wb.save(out_path)
    print(f"  Đã lưu -> {out_path}")


def main():
    parser = argparse.ArgumentParser(description="OCR cell images and write to Excel.")
    parser.add_argument("--input", default="Pages",
                        help="Folder containing page sub-folders (default: Pages)")
    parser.add_argument("--page", type=int, default=None,
                        help="Process only this page number; omit to process all pages")
    parser.add_argument("--bank", default="bidv",
                        help="Bank format to use — selects banks/<bank>.py (default: bidv)")
    args = parser.parse_args()

    bank_module = importlib.import_module(f"banks.{args.bank}")
    column_allowlists = getattr(bank_module, "COLUMN_ALLOWLISTS", {})
    print(f"Ngân hàng: {args.bank} — danh sách cho phép được định nghĩa cho các cột: {sorted(column_allowlists)}")

    reader = easyocr.Reader(["en"])

    if args.page:
        pages = [args.page]
    else:
        pages = sorted(
            int(p) for p in os.listdir(args.input)
            if os.path.isdir(os.path.join(args.input, p)) and p.isdigit()
        )

    for page_num in pages:
        page_dir = os.path.join(args.input, str(page_num))
        print(f"Đang xử lý trang {page_num}...")
        cells_to_excel(page_dir, reader, column_allowlists)


if __name__ == "__main__":
    main()
