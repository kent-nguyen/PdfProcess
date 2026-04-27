import os
import argparse
import importlib
from openpyxl import load_workbook


def find_raw_file(page_dir, page_num):
    path = os.path.join(page_dir, f"raw_{page_num}.xlsx")
    return path if os.path.exists(path) else None


def refine_page(raw_path, out_path, formatters, fixers):
    wb = load_workbook(raw_path)
    ws = wb.active

    row_errors = {}  # {row: {col: original_bad_value}} — populated by formatters, read by fixers
    row_fixes = {}   # {row: [note, ...]} — populated by fixers

    for formatter in formatters:
        formatter(ws, row_errors)

    for fixer in fixers:
        fixer(ws, row_fixes, row_errors, raw_path=raw_path)

    # Append Fixed / Error / Notes columns after the last data column
    fixed_col = ws.max_column + 1
    error_col = ws.max_column + 2
    notes_col = ws.max_column + 3
    ws.cell(row=1, column=fixed_col).value = "Fixed"
    ws.cell(row=1, column=error_col).value = "Error"
    ws.cell(row=1, column=notes_col).value = "Notes"

    for row in range(2, ws.max_row + 1):
        notes = []
        if row in row_fixes:
            ws.cell(row=row, column=fixed_col).value = True
            notes.extend(row_fixes[row])
        if row in row_errors:
            ws.cell(row=row, column=error_col).value = True
            notes.extend(row_errors[row].values())
        if notes:
            ws.cell(row=row, column=notes_col).value = "; ".join(str(n) for n in notes)

    wb.save(out_path)

    fixed_count = len(row_fixes)
    error_count = len(row_errors)
    print(f"  {fixed_count} dòng đã sửa, {error_count} dòng có lỗi -> {out_path}")

    return row_errors, row_fixes


def _print_summary(page_results):
    """Print a consolidated error/fix summary across all processed pages."""
    pages_with_issues = {p: (errs, fixes) for p, (errs, fixes) in page_results.items() if errs or fixes}
    if not pages_with_issues:
        print("\nKhông tìm thấy lỗi nào trên tất cả các trang.")
        return

    total_error_rows = sum(len(errs) for errs, _ in pages_with_issues.values())
    print(f"\n{'='*50}")
    print(f"Tóm tắt lỗi — {total_error_rows} dòng lỗi trên {len(pages_with_issues)} trang")
    print(f"{'='*50}")

    for page_num in sorted(pages_with_issues):
        row_errors, row_fixes = pages_with_issues[page_num]
        all_rows = sorted(set(row_errors) | set(row_fixes))
        print(f"\nTrang {page_num}: {len(all_rows)} dòng")
        for row in all_rows:
            notes = row_fixes.get(row, [])
            bad_cols = row_errors.get(row, {})
            if notes:
                for note in notes:
                    print(f"  Dòng {row}: {note}")
            if bad_cols:
                for col, original in bad_cols.items():
                    if not notes:  # format error with no fixer note
                        print(f"  Dòng {row}, cột {col}: lỗi định dạng (giá trị gốc: {original!r})")


def main():
    parser = argparse.ArgumentParser(description="Refine raw OCR Excel data column by column.")
    parser.add_argument("--input", default="Pages",
                        help="Folder containing page sub-folders (default: Pages)")
    parser.add_argument("--page", type=int, default=None,
                        help="Process only this page number; omit to process all pages")
    parser.add_argument("--bank", default="bidv",
                        help="Bank format to use — selects fixers/<bank>.py (default: bidv)")
    args = parser.parse_args()

    fixer_module = importlib.import_module(f"banks.{args.bank}")
    formatters = getattr(fixer_module, "FORMATTERS", [])
    fixers = getattr(fixer_module, "FIXERS", [])
    print(f"Ngân hàng: {args.bank} — đã tải {len(formatters)} bộ định dạng, {len(fixers)} bộ sửa lỗi")

    if args.page:
        pages = [args.page]
    else:
        pages = sorted(
            int(p) for p in os.listdir(args.input)
            if os.path.isdir(os.path.join(args.input, p)) and p.isdigit()
        )

    page_results = {}  # {page_num: (row_errors, row_fixes)}

    for page_num in pages:
        page_dir = os.path.join(args.input, str(page_num))
        raw_path = find_raw_file(page_dir, page_num)
        if not raw_path:
            print(f"Không tìm thấy file Excel thô trong {page_dir}, bỏ qua.")
            continue
        out_path = os.path.join(page_dir, f"{page_num}.xlsx")
        print(f"Đang xử lý trang {page_num} ({os.path.basename(raw_path)})...")
        page_results[page_num] = refine_page(raw_path, out_path, formatters, fixers)

    _print_summary(page_results)


if __name__ == "__main__":
    main()
